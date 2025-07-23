from http.server import BaseHTTPRequestHandler
import json
import urllib.parse
import tempfile
import os
from datetime import datetime

class handler(BaseHTTPRequestHandler):
    
    def parse_multipart(self, data, boundary):
        """Parse multipart form data para extrair arquivos XML"""
        files = []
        try:
            boundary_bytes = boundary.encode('utf-8')
            parts = data.split(b'--' + boundary_bytes)
            
            for part in parts:
                if b'Content-Disposition' in part and b'filename=' in part:
                    lines = part.split(b'\r\n')
                    filename = None
                    content_start = -1
                    
                    for i, line in enumerate(lines):
                        if b'Content-Disposition' in line:
                            line_str = line.decode('utf-8', errors='ignore')
                            if 'filename=' in line_str:
                                start = line_str.find('filename="') + 10
                                end = line_str.find('"', start)
                                filename = line_str[start:end] if start > 9 and end > start else None
                        elif line == b'' and content_start == -1:
                            content_start = i + 1
                            break
                    
                    if content_start > 0 and filename and filename.lower().endswith('.xml'):
                        content_lines = lines[content_start:]
                        if content_lines and content_lines[-1] == b'':
                            content_lines = content_lines[:-1]
                        file_content = b'\r\n'.join(content_lines)
                        if file_content:
                            files.append({'filename': filename, 'content': file_content})
        except:
            pass
        return files
    
    def processar_xml(self, content):
        """Processa um arquivo XML de NFe"""
        try:
            import xmltodict
            xml_dict = xmltodict.parse(content.decode('utf-8'))
            
            # Encontrar NFe
            if 'nfeProc' in xml_dict:
                nfe = xml_dict['nfeProc']['NFe']['infNFe']
            elif 'NFe' in xml_dict:
                nfe = xml_dict['NFe']['infNFe']
            else:
                return []
            
            ide = nfe.get('ide', {})
            emit = nfe.get('emit', {})
            total = nfe.get('total', {})
            
            # Dados base da nota
            base_data = {
                'numero_nf': ide.get('nNF', 'N/A'),
                'serie': ide.get('serie', '1'),
                'data_emissao': ide.get('dhEmi', ide.get('dEmi', datetime.now().strftime('%Y-%m-%d'))),
                'emitente': emit.get('xNome', 'N/A'),
                'cnpj_emitente': emit.get('CNPJ', 'N/A'),
                'valor_total_nf': total.get('ICMSTot', {}).get('vNF', '0')
            }
            
            # Processar itens
            det = nfe.get('det', [])
            if not isinstance(det, list):
                det = [det]
            
            result = []
            for item in det:
                prod = item.get('prod', {})
                item_data = base_data.copy()
                item_data.update({
                    'descricao_produto': prod.get('xProd', 'Produto'),
                    'quantidade_comercial': prod.get('qCom', '1'),
                    'valor_unitario': prod.get('vUnCom', '0'),
                    'valor_total_item': prod.get('vProd', '0'),
                    'cfop': prod.get('CFOP', '5102')
                })
                result.append(item_data)
            
            return result or [base_data]
            
        except Exception as e:
            return [{
                'numero_nf': 'ERRO',
                'data_emissao': datetime.now().strftime('%Y-%m-%d'),
                'emitente': f'Erro: {str(e)[:50]}',
                'erro': str(e)
            }]
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {"status": "ok", "message": "Datarum API Online", "timestamp": datetime.now().isoformat()}
        self.wfile.write(json.dumps(response).encode())
    
    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length) if content_length > 0 else b''
        
        if 'processar-info' in self.path:
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', '*')
            self.end_headers()
            
            try:
                # Extrair boundary do Content-Type
                content_type = self.headers.get('Content-Type', '')
                boundary = None
                if 'multipart/form-data' in content_type:
                    boundary_start = content_type.find('boundary=')
                    if boundary_start != -1:
                        boundary = content_type[boundary_start + 9:].strip()
                
                todos_dados = []
                stats = {'processados': 0, 'itens': 0, 'notas': [], 'emitentes': set(), 'valor_total': 0.0}
                
                if boundary and post_data:
                    arquivos_xml = self.parse_multipart(post_data, boundary)
                    
                    for arquivo in arquivos_xml:
                        dados_nfe = self.processar_xml(arquivo['content'])
                        if dados_nfe:
                            todos_dados.extend(dados_nfe)
                            stats['processados'] += 1
                            stats['itens'] += len(dados_nfe)
                            
                            for item in dados_nfe:
                                if item.get('numero_nf'):
                                    stats['notas'].append(item['numero_nf'])
                                if item.get('emitente'):
                                    stats['emitentes'].add(item['emitente'])
                                if item.get('valor_total_item'):
                                    try:
                                        valor = float(str(item['valor_total_item']).replace(',', '.'))
                                        stats['valor_total'] += valor
                                    except:
                                        pass
                
                # Salvar no cache global (simplificado)
                global cache_dados
                cache_dados = todos_dados
                
                response = {
                    'message': 'Processamento concluído!',
                    'arquivos_processados': stats['processados'],
                    'itens_processados': stats['itens'],
                    'notas_encontradas': list(stats['notas'])[:10],
                    'emitentes': list(stats['emitentes'])[:10],
                    'valor_total': round(stats['valor_total'], 2),
                    'session_id': f"session_{stats['processados']}",
                    'planilha_destino': 'datarum_processamento.xlsx'
                }
                
            except Exception as e:
                response = {'message': f'Erro: {str(e)}', 'erro': True}
            
            self.wfile.write(json.dumps(response).encode())
        
        elif 'processar' in self.path:
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Disposition', 'attachment; filename=datarum_processamento.xlsx')
            self.end_headers()
            
            try:
                import tempfile
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                # Buscar dados do cache
                global cache_dados
                dados = cache_dados if 'cache_dados' in globals() else []
                
                if not dados:
                    # Dados de exemplo se não houver cache
                    dados = [{
                        'numero_nf': '123456',
                        'serie': '1',
                        'data_emissao': datetime.now().strftime('%Y-%m-%d'),
                        'emitente': 'Empresa Exemplo Ltda',
                        'descricao_produto': 'Produto de Exemplo',
                        'quantidade_comercial': '10',
                        'valor_unitario': '100.00',
                        'valor_total_item': '1000.00',
                        'cfop': '5102'
                    }]
                
                # Criar workbook Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "NFe Processadas"
                
                # Headers com formatação
                headers = ['Numero NF', 'Serie', 'Data Emissao', 'Emitente', 'CNPJ Emitente', 
                          'Produto', 'Quantidade', 'Valor Unitario', 'Total Item', 'CFOP']
                
                # Estilo do cabeçalho
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                # Adicionar cabeçalhos
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Adicionar dados
                for row_idx, item in enumerate(dados, 2):
                    ws.cell(row=row_idx, column=1, value=item.get('numero_nf', ''))
                    ws.cell(row=row_idx, column=2, value=item.get('serie', ''))
                    ws.cell(row=row_idx, column=3, value=item.get('data_emissao', ''))
                    ws.cell(row=row_idx, column=4, value=item.get('emitente', ''))
                    ws.cell(row=row_idx, column=5, value=item.get('cnpj_emitente', ''))
                    ws.cell(row=row_idx, column=6, value=item.get('descricao_produto', ''))
                    ws.cell(row=row_idx, column=7, value=item.get('quantidade_comercial', ''))
                    ws.cell(row=row_idx, column=8, value=item.get('valor_unitario', ''))
                    ws.cell(row=row_idx, column=9, value=item.get('valor_total_item', ''))
                    ws.cell(row=row_idx, column=10, value=item.get('cfop', ''))
                
                # Ajustar largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Salvar em arquivo temporário
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    wb.save(tmp_file.name)
                    
                    # Ler arquivo e enviar
                    with open(tmp_file.name, 'rb') as f:
                        excel_content = f.read()
                    
                    # Limpar arquivo temporário
                    os.unlink(tmp_file.name)
                
                self.wfile.write(excel_content)
                
            except Exception as e:
                # Em caso de erro, fallback para CSV
                self.send_response(200)
                self.send_header('Content-Type', 'text/csv; charset=utf-8')
                self.send_header('Content-Disposition', 'attachment; filename=datarum_erro.csv')
                
                error_csv = f'Erro,{str(e)}\nContate o suporte,suporte@datarum.com.br'
                self.wfile.write(error_csv.encode('utf-8'))
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', '*')
        self.send_header('Access-Control-Allow-Headers', '*')
        self.end_headers()

# Cache global simples
cache_dados = []
