import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from api.config import CAMPOS_DISPONIVEIS


def salvar_em_excel(dados: List[Dict], caminho_planilha: str, configuracao: Dict[str, Any] = None):
    if not dados:
        return 0
    
    config = configuracao or {}
    opcoes = config.get('opcoes', {})
    preset = config.get('preset', 'basico')

    df = pd.DataFrame(dados)

    # Renomeação dinâmica
    mapeamento_labels = {
        k: v['label']
        for k, v in CAMPOS_DISPONIVEIS.items()
        if k in df.columns
    }
    df.rename(columns=mapeamento_labels, inplace=True)

    # Conversão numérica
    keywords_numericas = ["Vl.", "Total", "Base", "Aliq.", "Qtd", "Desconto"]
    cols_para_converter = [
        col for col in df.columns
        if any(key in col for key in keywords_numericas)
    ]

    for col in cols_para_converter:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    caminho = Path(caminho_planilha)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        nome_aba = f"Dados {preset.title()}"
        df.to_excel(writer, sheet_name=nome_aba, index=False)

        workbook = writer.book
        worksheet = writer.sheets[nome_aba]

        # Estilo do Cabeçalho
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto ajuste correto de colunas
        for idx, col in enumerate(df.columns, 1):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2

            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max_len, 500)

        # Filtro automático
        worksheet.auto_filter.ref = worksheet.dimensions

        # Congelar cabeçalho
        worksheet.freeze_panes = "A2"

        # Aba de Resumo
        if opcoes.get('incluirResumo'):
            label_total = CAMPOS_DISPONIVEIS.get("valor_total_item", {}).get("label")
            total_fin = df[label_total].sum() if label_total in df.columns else 0

            df_resumo = pd.DataFrame({
                "Métrica": [
                    "Total de Linhas (Itens)",
                    "Valor Bruto Total",
                    "Data do Processamento"
                ],
                "Valor": [
                    len(df),
                    total_fin,
                    pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
                ]
            })

            df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

    return len(df)